# -*- coding: utf-8 -*-
"""
Aplica prisma/sql/006_dm_cr.sql e importa a planilha "Centro_de_Custo" -> dm_cr.
Só as 16 colunas de cabecalho VERDE. CR normalizado para 5 caracteres.

Uso:
    python scripts/importar_dm_cr.py "C:/caminho/Centro_de_Custo.xlsx"

Idempotente: faz upsert por CR (reimportar atualiza os registros).
Le DATABASE_URL de .env.local. Requer: openpyxl, psycopg2.
"""
import os
import re
import sys

import openpyxl
import psycopg2
import psycopg2.extras

PROJ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# cabecalho da planilha -> coluna da tabela (apenas as verdes)
MAP = {
    "DATA INICIO CR": "data_inicio_cr",
    "BLOQUEIO": "bloqueio",
    "CR": "cr",
    "DESCRI CR": "descri_cr",
    "REGIONAL": "regional",
    "CONQ PERD": "conq_perd",
    "DESCRI NEGOCIO": "descri_negocio",
    "DESCRI SOLUCAO": "descri_solucao",
    "NOME CLIENTE": "nome_cliente",
    "NOME GRP CLIENTE": "nome_grp_cliente",
    "PEC": "pec",
    "DIRETOR EXECUTIVO": "diretor_executivo",
    "DIRETOR REGIONAL": "diretor_regional",
    "GERENTE REGIONAL": "gerente_regional",
    "GERENTE": "gerente",
    "SUPERVISOR": "supervisor",
}
COLS = list(MAP.values())


def ler_env(nome):
    with open(os.path.join(PROJ, ".env.local"), "r", encoding="utf-8") as f:
        for linha in f:
            m = re.match(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.*?)\s*$", linha)
            if m and m.group(1) == nome:
                v = m.group(2)
                if len(v) >= 2 and v[0] == v[-1] and v[0] in ("'", '"'):
                    v = v[1:-1]
                return v
    raise RuntimeError(f"{nome} nao encontrado em .env.local")


def normalizar_cr(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    if len(s) < 5:
        s = s.zfill(5)  # "1489" -> "01489"; sempre 5 caracteres
    return s


def parse_data(v):
    if v is None:
        return None
    s = str(v).strip()
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if not m:
        return None
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"  # ISO


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main():
    if len(sys.argv) < 2:
        print('Informe o caminho da planilha. Ex.: python scripts/importar_dm_cr.py "C:/.../Centro_de_Custo.xlsx"')
        sys.exit(1)
    xlsx_path = sys.argv[1]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Centro_de_Custo"] if "Centro_de_Custo" in wb.sheetnames else wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {}
    for hdr, col in MAP.items():
        if hdr not in header:
            raise RuntimeError(f"Cabecalho nao encontrado na planilha: {hdr}")
        idx[col] = header.index(hdr)

    registros = []
    vistos = set()
    sem_cr = colisao = padded = 0
    for linha in rows:
        cr_raw = linha[idx["cr"]]
        cr = normalizar_cr(cr_raw)
        if not cr:
            sem_cr += 1
            continue
        if cr_raw is not None and len(str(cr_raw).strip()) < 5:
            padded += 1
        if cr in vistos:
            colisao += 1
            continue
        vistos.add(cr)
        reg = []
        for col in COLS:
            val = linha[idx[col]]
            if col == "cr":
                reg.append(cr)
            elif col == "data_inicio_cr":
                reg.append(parse_data(val))
            else:
                reg.append(txt(val))
        registros.append(tuple(reg))

    print(f"Registros validos: {len(registros)} | sem CR: {sem_cr} | zero-padeados: {padded} | colisoes ignoradas: {colisao}")

    conn = psycopg2.connect(ler_env("DATABASE_URL"))
    try:
        with conn.cursor() as cur:
            with open(os.path.join(PROJ, "prisma", "sql", "006_dm_cr.sql"), "r", encoding="utf-8") as f:
                cur.execute(f.read())
            print("SQL 006_dm_cr.sql aplicado (tabela criada/garantida).")

            col_list = ", ".join(f'"{c}"' for c in COLS)
            update_set = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in COLS if c != "cr")
            sql = (
                f'INSERT INTO "dm_cr" ({col_list}) VALUES %s '
                f'ON CONFLICT ("cr") DO UPDATE SET {update_set}, "atualizado_em" = now()'
            )
            psycopg2.extras.execute_values(cur, sql, registros, page_size=500)
            conn.commit()
            print(f"Upsert concluido: {len(registros)} registros.")

            cur.execute('SELECT count(*) FROM "dm_cr"')
            print("Total na tabela dm_cr:", cur.fetchone()[0])
            cur.execute("SELECT cr, nome_grp_cliente, regional FROM \"dm_cr\" WHERE cr LIKE '0%' ORDER BY cr LIMIT 5")
            print("CRs zero-padeados (amostra):", cur.fetchall())
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
